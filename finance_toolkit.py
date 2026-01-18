import tkinter as tk
from tkinter import filedialog
from tkinter import ttk, messagebox
import csv
import os, datetime

# ---------- Optional libraries ----------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except Exception:
    MATPLOTLIB_AVAILABLE = False

# ---------- Theme ----------
PRIMARY_BG = "#1D3557"   # Main screen background
SIDEBAR_BG = "#1A2C4A"   # Sidebar + buttons
HEADING_FG = "#F1FA8C"   # Headings
TEXT_FG = "white"        # Normal labels
ACCENT_BTN = "#06D6A0"   # Used for some action buttons
ACCENT_BTN_2 = "#FFD166" # Secondary action buttons
ACCENT_LINE = "#A5D8FF"  # Chart lines/bars
TABLE_HEADER_FILL = "2563EB"

# ---------- Folders ----------
REPORTS_DIR = "FinanceReports"
os.makedirs(REPORTS_DIR, exist_ok=True)

# ---------- Helpers ----------
def format_currency(x):
    try:
        return f"₹{float(x):,.2f}"
    except Exception:
        return x

def today_str():
    return datetime.date.today().strftime("%Y-%m-%d")

def save_to_excel_or_csv(default_name, headers, rows):
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx" if OPENPYXL_AVAILABLE else ".csv",
        initialfile=default_name,
        filetypes=[("Excel files","*.xlsx"),("CSV files","*.csv")]
    )
    if not filename:
        return None
    if OPENPYXL_AVAILABLE and filename.endswith(".xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append(r)
        bold = Font(bold=True)
        fill = PatternFill(start_color=TABLE_HEADER_FILL, end_color=TABLE_HEADER_FILL, fill_type="solid")
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for i, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 4
        wb.save(filename)
    else:
        with open(filename, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
    return filename

# ---------- Expense Tracker ----------
def show_expense_tracker(frame):
    for w in frame.winfo_children(): w.destroy()
    bg = PRIMARY_BG
    frame.configure(bg=bg)

    tk.Label(frame, text="Expense Tracker", font=("Segoe UI", 18, "bold"),
             fg=HEADING_FG, bg=bg).pack(pady=(10, 8))

    form = tk.Frame(frame, bg=bg)
    form.pack(fill="x", padx=16)
    tk.Label(form, text="Date (YYYY-MM-DD):", bg=bg, fg=TEXT_FG).grid(row=0, column=0, sticky="w", padx=4, pady=4)
    entry_date = tk.Entry(form)
    entry_date.grid(row=0, column=1, padx=4, pady=4)

    tk.Label(form, text="Type:", bg=bg, fg=TEXT_FG).grid(row=0, column=2, sticky="w", padx=4, pady=4)
    combo_type = ttk.Combobox(form, values=["Expense","Income"], state="readonly", width=12)
    combo_type.grid(row=0, column=3, padx=4, pady=4)

    tk.Label(form, text="Category:", bg=bg, fg=TEXT_FG).grid(row=1, column=0, sticky="w", padx=4, pady=4)
    default_categories = ["Food","Rent","Transport","Utilities","Entertainment","Salary","Investment","Other"]
    combo_category = ttk.Combobox(form, values=default_categories, width=20)
    combo_category.grid(row=1, column=1, padx=4, pady=4)

    tk.Label(form, text="Amount (₹):", bg=bg, fg=TEXT_FG).grid(row=1, column=2, sticky="w", padx=4, pady=4)
    entry_amount = tk.Entry(form)
    entry_amount.grid(row=1, column=3, padx=4, pady=4)

    tk.Label(form, text="Note:", bg=bg, fg=TEXT_FG).grid(row=2, column=0, sticky="w", padx=4, pady=4)
    entry_note = tk.Entry(form, width=60)
    entry_note.grid(row=2, column=1, columnspan=3, padx=4, pady=4, sticky="w")

    btn_frame = tk.Frame(frame, bg=bg); btn_frame.pack(fill="x", padx=16, pady=(6,8))

    if not hasattr(show_expense_tracker, "entries"):
        show_expense_tracker.entries = []

    def update_summary():
        income = sum(e["amount"] for e in show_expense_tracker.entries if e["type"]=="Income")
        expense = sum(e["amount"] for e in show_expense_tracker.entries if e["type"]=="Expense")
        balance = income - expense
        summary_var.set(
            f"Income: {format_currency(income)}   |   "
            f"Expenses: {format_currency(expense)}   |   "
            f"Balance: {format_currency(balance)}"
        )

    def refresh_table():
        for r in tree.get_children(): tree.delete(r)
        for idx,e in enumerate(show_expense_tracker.entries):
            tree.insert("", "end", iid=str(idx),
                        values=(e["date"],e["type"],e["category"],
                                format_currency(e["amount"]),e["note"]))
        update_summary()

    def add_entry():
        d = entry_date.get().strip()
        t = combo_type.get().strip()
        cat = combo_category.get().strip() or "Other"
        note = entry_note.get().strip()
        if not t:
            messagebox.showerror("Invalid", "Select type (Expense/Income)."); return
        try:
            datetime.datetime.strptime(d, "%Y-%m-%d")
        except Exception:
            messagebox.showerror("Invalid date", "Use YYYY-MM-DD."); return
        try:
            amt = float(entry_amount.get().strip())
            if amt <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Invalid amount", "Enter positive numeric amount."); return
        show_expense_tracker.entries.append(
            {"date":d,"type":t,"category":cat,"amount":amt,"note":note}
        )
        entry_amount.delete(0,tk.END); entry_note.delete(0,tk.END)
        refresh_table()

    def delete_selected():
        sel = tree.selection()
        if not sel: return
        for item in sorted(sel, reverse=True):
            idx = int(item); show_expense_tracker.entries.pop(idx)
        refresh_table()

    def export_entries():
        if not show_expense_tracker.entries:
            messagebox.showinfo("No data","Nothing to export."); return
        headers = ["Date","Type","Category","Amount","Note"]
        rows = [[e["date"],e["type"],e["category"],e["amount"],e["note"]]
                for e in show_expense_tracker.entries]
        path = save_to_excel_or_csv(
            os.path.join(REPORTS_DIR,f"Expenses_{today_str()}"),
            headers, rows
        )
        if path: messagebox.showinfo("Exported", f"Saved to {path}")

    ttk.Button(btn_frame, text="Add Entry", command=add_entry).pack(side="left", padx=6)
    ttk.Button(btn_frame, text="Delete Selected", command=delete_selected).pack(side="left", padx=6)
    ttk.Button(btn_frame, text="Export", command=export_entries).pack(side="left", padx=6)

    tree_frame = tk.Frame(frame, bg=bg); tree_frame.pack(fill="both", expand=True, padx=16, pady=(6,12))
    cols = ("date","type","category","amount","note")
    tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=12)
    for c in cols: tree.heading(c, text=c.capitalize())
    tree.column("date", width=100, anchor="center")
    tree.column("type", width=80, anchor="center")
    tree.column("category", width=120, anchor="center")
    tree.column("amount", width=110, anchor="e")
    tree.column("note", width=260, anchor="w")
    vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscroll=vsb.set); vsb.pack(side="right", fill="y"); tree.pack(side="left", fill="both", expand=True)

    summary_var = tk.StringVar()
    summary_var.set("Income: ₹0.00   |   Expenses: ₹0.00   |   Balance: ₹0.00")
    tk.Label(frame, textvariable=summary_var, bg=SIDEBAR_BG, fg="#A5D8FF",
             pady=6, font=("Segoe UI",10)).pack(fill="x")

    refresh_table()

# ---------- Step-up vs Normal SIP ----------
def show_step_up_vs_sip(frame):
    for w in frame.winfo_children(): w.destroy()
    bg = PRIMARY_BG
    frame.configure(bg=bg)

    tk.Label(frame, text="Step-Up SIP vs Normal SIP", font=("Segoe UI", 18, "bold"),
             fg=HEADING_FG, bg=bg).pack(pady=(10,8))

    grid = tk.Frame(frame, bg=bg); grid.pack(fill="x", padx=16, pady=(0,8))
    labels = ["Monthly SIP (₹)","Duration (years)","Expected Annual Return (%)",
              "Step-Up % per year","Expected Inflation (%)"]
    entries = {}
    for i,lbl in enumerate(labels):
        tk.Label(grid, text=lbl, bg=bg, fg=TEXT_FG).grid(row=i, column=0, sticky="w", padx=6, pady=6)
        e = tk.Entry(grid, width=14); e.grid(row=i, column=1, sticky="w", padx=6, pady=6)
        entries[lbl] = e

    btn_row = tk.Frame(frame, bg=bg); btn_row.pack(fill="x", padx=16, pady=(0,8))
    compare_btn = tk.Button(btn_row, text="Compare SIPs", bg=ACCENT_BTN, fg="black",
                            font=("Segoe UI",10,"bold"))
    export_btn = tk.Button(btn_row, text="Export Results", bg=ACCENT_BTN_2, fg="black",
                           state="disabled", font=("Segoe UI",10,"bold"))
    compare_btn.pack(side="left", padx=(0,8)); export_btn.pack(side="left", padx=8)

    table_frame = tk.Frame(frame, bg=bg); table_frame.pack(fill="both", expand=False, padx=16, pady=(6,8))
    cols = ("year","step_monthly","step_invested","step_fv","norm_monthly",
            "norm_invested","norm_fv","inflation_adj_step","inflation_adj_norm")
    tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=8)
    headings = {
        "year":"Year","step_monthly":"Step-up ₹/mo","step_invested":"Invested Step-up",
        "step_fv":"FV Step-up","norm_monthly":"Normal ₹/mo","norm_invested":"Invested Normal",
        "norm_fv":"FV Normal","inflation_adj_step":"InflAdj Step","inflation_adj_norm":"InflAdj Norm"
    }
    for c in cols: tree.heading(c, text=headings[c])
    tree.column("year", width=50, anchor="center")
    for c in cols[1:]: tree.column(c, width=110, anchor="e")
    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscroll=vsb.set); vsb.pack(side="right", fill="y"); tree.pack(side="left", fill="both", expand=True)

    result_frame = tk.Frame(frame, bg=SIDEBAR_BG); result_frame.pack(fill="x", padx=16, pady=(6,8))
    result_label = tk.Label(result_frame, text="", bg=SIDEBAR_BG, fg="#E5FBFF",
                            font=("Segoe UI",10), pady=6, justify="left", wraplength=1100)
    result_label.pack(fill="x")

    chart_frame = tk.Frame(frame, bg=bg); chart_frame.pack(fill="both", expand=True, padx=16, pady=(6,12))
    chart_canvas_container = {"canvas": None}

    def clear_chart():
        c = chart_canvas_container.get("canvas")
        if c:
            c.get_tk_widget().destroy()
            chart_canvas_container["canvas"] = None

    def calculate_and_display():
        try:
            sip = float(entries["Monthly SIP (₹)"].get())
            years = int(entries["Duration (years)"].get())
            rate = float(entries["Expected Annual Return (%)"].get())
            step_up = float(entries["Step-Up % per year"].get())
            inflation = float(entries["Expected Inflation (%)"].get())
            if years <= 0 or sip <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Invalid input","Enter positive numbers in all fields."); return

        monthly_rate = rate/12/100
        total_months = years*12

        invested_step_by_year = [0.0]*years
        fv_step_by_year = [0.0]*years
        invested_norm_by_year = [0.0]*years
        fv_norm_by_year = [0.0]*years

        fv_step = fv_norm = 0.0
        invested_step_total = invested_norm_total = 0.0

        for m in range(total_months):
            y = m//12
            monthly_step = sip * ((1+step_up/100)**y)
            fv_step = fv_step*(1+monthly_rate) + monthly_step
            invested_step_total += monthly_step
            invested_step_by_year[y] += monthly_step

            fv_norm = fv_norm*(1+monthly_rate) + sip
            invested_norm_total += sip
            invested_norm_by_year[y] += sip

            if (m+1)%12 == 0:
                fv_step_by_year[y] = fv_step
                fv_norm_by_year[y] = fv_norm

        inflation_adj_step_total = fv_step / ((1+inflation/100)**years)
        inflation_adj_norm_total = fv_norm / ((1+inflation/100)**years)
        inflation_adj_step_by_year = [
            fv_step_by_year[i]/((1+inflation/100)**(years-i-1)) for i in range(years)
        ]
        inflation_adj_norm_by_year = [
            fv_norm_by_year[i]/((1+inflation/100)**(years-i-1)) for i in range(years)
        ]

        for r in tree.get_children(): tree.delete(r)
        for i in range(years):
            tree.insert("", "end", values=(
                i+1,
                format_currency(round(sip*((1+step_up/100)**i),2)),
                format_currency(round(invested_step_by_year[i],2)),
                format_currency(round(fv_step_by_year[i],2)),
                format_currency(round(sip,2)),
                format_currency(round(invested_norm_by_year[i],2)),
                format_currency(round(fv_norm_by_year[i],2)),
                format_currency(round(inflation_adj_step_by_year[i],2)),
                format_currency(round(inflation_adj_norm_by_year[i],2)),
            ))

        diff = fv_step - fv_norm
        result_label.config(text=(
            f"Total Invested → Step-up: {format_currency(invested_step_total)} | "
            f"Normal: {format_currency(invested_norm_total)}    "
            f"Final FV → Step-up: {format_currency(fv_step)} | "
            f"Normal: {format_currency(fv_norm)}    "
            f"Difference: {format_currency(diff)}    "
            f"Inflation-adjusted → Step-up: {format_currency(inflation_adj_step_total)} | "
            f"Normal: {format_currency(inflation_adj_norm_total)}"
        ))

        export_btn.config(state="normal")

        clear_chart()
        if not MATPLOTLIB_AVAILABLE:
            tk.Label(chart_frame, text="Install matplotlib for chart.", fg=TEXT_FG, bg=bg).pack()
            return

        years_list = list(range(1, years+1))
        fig = Figure(figsize=(9,3.2), dpi=95)
        ax = fig.add_subplot(111)
        width = 0.35
        ax.bar([y-width/2 for y in years_list], fv_norm_by_year, width=width, label="Normal SIP", color=ACCENT_LINE)
        ax.bar([y+width/2 for y in years_list], fv_step_by_year, width=width, label="Step-up SIP", color=ACCENT_BTN)
        ax.set_facecolor(bg); fig.patch.set_facecolor(bg)
        ax.set_xlabel("Year"); ax.set_ylabel("Future Value (₹)")
        ax.set_title("Year-by-Year FV: Normal vs Step-up", color=HEADING_FG)
        ax.tick_params(colors=TEXT_FG)
        for spine in ax.spines.values(): spine.set_color(TEXT_FG)
        ax.xaxis.label.set_color(TEXT_FG); ax.yaxis.label.set_color(TEXT_FG)
        ax.legend(facecolor=SIDEBAR_BG, edgecolor=TEXT_FG)
        ax.grid(axis="y", linestyle="--", alpha=0.3, color=ACCENT_LINE)
        canvas = FigureCanvasTkAgg(fig, master=chart_frame); canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        chart_canvas_container["canvas"] = canvas

        tree._calc = {
            "years": years, "sip": sip, "step_up": step_up,
            "invested_step_by_year": invested_step_by_year,
            "fv_step_by_year": fv_step_by_year,
            "invested_norm_by_year": invested_norm_by_year,
            "fv_norm_by_year": fv_norm_by_year,
            "invested_step_total": invested_step_total,
            "invested_norm_total": invested_norm_total,
            "fv_step_total": fv_step,
            "fv_norm_total": fv_norm,
            "inflation_adj_step_total": inflation_adj_step_total,
            "inflation_adj_norm_total": inflation_adj_norm_total,
            "inflation_adj_step_by_year": inflation_adj_step_by_year,
            "inflation_adj_norm_by_year": inflation_adj_norm_by_year
        }

    def export_comparison():
        data = getattr(tree, "_calc", None)
        if not data:
            messagebox.showinfo("No data","Run comparison first."); return
        headers = [
            "Year","Step-up Monthly","Invested Step-up","FV Step-up",
            "Normal Monthly","Invested Normal","FV Normal","InflAdj Step","InflAdj Normal"
        ]
        rows = []
        for i in range(data["years"]):
            rows.append([
                i+1,
                round(data["sip"]*((1+data["step_up"]/100)**i),2),
                round(data["invested_step_by_year"][i],2),
                round(data["fv_step_by_year"][i],2),
                round(data["sip"],2),
                round(data["invested_norm_by_year"][i],2),
                round(data["fv_norm_by_year"][i],2),
                round(data["inflation_adj_step_by_year"][i],2),
                round(data["inflation_adj_norm_by_year"][i],2)
            ])
        rows.append([])
        rows.append(["Total Invested (Step-up)", data["invested_step_total"]])
        rows.append(["Total Invested (Normal)", data["invested_norm_total"]])
        rows.append(["Final FV (Step-up)", data["fv_step_total"]])
        rows.append(["Final FV (Normal)", data["fv_norm_total"]])
        rows.append(["Inflation-adjusted (Step-up)", data["inflation_adj_step_total"]])
        rows.append(["Inflation-adjusted (Normal)", data["inflation_adj_norm_total"]])
        path = save_to_excel_or_csv(
            os.path.join(REPORTS_DIR,f"StepUp_vs_SIP_{today_str()}"),
            headers, rows
        )
        if path: messagebox.showinfo("Exported", f"Saved to {path}")

    compare_btn.config(command=calculate_and_display)
    export_btn.config(command=export_comparison)

# ---------- SIP Calculator ----------
def show_sip_calculator(frame):
    for w in frame.winfo_children(): w.destroy()
    bg = PRIMARY_BG
    frame.configure(bg=bg)

    tk.Label(frame, text="SIP Calculator", font=("Segoe UI", 18, "bold"),
             fg=HEADING_FG, bg=bg).pack(pady=(10,8))

    grid = tk.Frame(frame, bg=bg); grid.pack(fill="x", padx=16, pady=(0,8))
    labels = ["Monthly SIP (₹)","Duration (years)","Expected Annual Return (%)","Expected Inflation (%)"]
    entries = {}
    for i,lbl in enumerate(labels):
        tk.Label(grid, text=lbl, bg=bg, fg=TEXT_FG).grid(row=i, column=0, sticky="w", padx=6, pady=6)
        e = tk.Entry(grid, width=16); e.grid(row=i, column=1, sticky="w", padx=6, pady=6)
        entries[lbl] = e

    btn_row = tk.Frame(frame, bg=bg); btn_row.pack(fill="x", padx=16, pady=(0,8))
    calc_btn = tk.Button(btn_row, text="Calculate SIP", bg=ACCENT_BTN, fg="black",
                         font=("Segoe UI",10,"bold"))
    export_btn = tk.Button(btn_row, text="Export Results", bg=ACCENT_BTN_2, fg="black",
                           state="disabled", font=("Segoe UI",10,"bold"))
    calc_btn.pack(side="left", padx=(0,8)); export_btn.pack(side="left", padx=8)

    table_frame = tk.Frame(frame, bg=bg); table_frame.pack(fill="both", expand=False, padx=16, pady=(6,8))
    cols = ("year","monthly","invested","fv","inflation_adj")
    tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=8)
    heads = {"year":"Year","monthly":"Monthly ₹/mo","invested":"Invested ₹","fv":"FV ₹","inflation_adj":"InflAdj FV ₹"}
    for c in cols: tree.heading(c, text=heads[c])
    tree.column("year", width=60, anchor="center")
    for c in cols[1:]: tree.column(c, width=130, anchor="e")
    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscroll=vsb.set); vsb.pack(side="right", fill="y"); tree.pack(side="left", fill="both", expand=True)

    result_frame = tk.Frame(frame, bg=SIDEBAR_BG); result_frame.pack(fill="x", padx=16, pady=(6,8))
    result_label = tk.Label(result_frame, text="", bg=SIDEBAR_BG, fg="#BBF7D0",
                            font=("Segoe UI",10), pady=6, justify="left", wraplength=1100)
    result_label.pack(fill="x")

    chart_frame = tk.Frame(frame, bg=bg); chart_frame.pack(fill="both", expand=True, padx=16, pady=(6,12))
    chart_canvas_container = {"canvas": None}

    def clear_chart():
        c = chart_canvas_container.get("canvas")
        if c:
            c.get_tk_widget().destroy()
            chart_canvas_container["canvas"] = None

    def calculate_sip():
        try:
            sip = float(entries["Monthly SIP (₹)"].get())
            years = int(entries["Duration (years)"].get())
            rate = float(entries["Expected Annual Return (%)"].get())
            inflation = float(entries["Expected Inflation (%)"].get())
            if years <= 0 or sip <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Invalid input","Enter positive numbers in all fields."); return

        monthly_rate = rate/12/100
        total_months = years*12
        fv = 0.0
        invested_by_year = [0.0]*years
        fv_by_year = [0.0]*years
        invested_total = 0.0

        for m in range(total_months):
            y = m//12
            fv = fv*(1+monthly_rate) + sip
            invested_by_year[y] += sip
            invested_total += sip
            if (m+1)%12 == 0: fv_by_year[y] = fv

        inflation_adj_total = fv / ((1+inflation/100)**years)
        inflation_adj_by_year = [
            fv_by_year[i]/((1+inflation/100)**(years-i-1)) for i in range(years)
        ]

        for r in tree.get_children(): tree.delete(r)
        for i in range(years):
            tree.insert("", "end", values=(
                i+1,
                format_currency(sip),
                format_currency(round(invested_by_year[i],2)),
                format_currency(round(fv_by_year[i],2)),
                format_currency(round(inflation_adj_by_year[i],2))
            ))

        result_label.config(text=(
            f"Total Invested: {format_currency(invested_total)}    "
            f"Final FV: {format_currency(fv)}    "
            f"Inflation-adjusted FV: {format_currency(inflation_adj_total)}    "
            f"Profit: {format_currency(fv - invested_total)}"
        ))

        export_btn.config(state="normal")

        clear_chart()
        if not MATPLOTLIB_AVAILABLE:
            tk.Label(chart_frame, text="Install matplotlib for chart.", fg=TEXT_FG, bg=bg).pack()
            tree._calc = {
                "years": years,
                "sip": sip,
                "invested_by_year": invested_by_year,
                "fv_by_year": fv_by_year,
                "invested_total": invested_total,
                "fv_total": fv,
                "inflation_adj_total": inflation_adj_total,
                "inflation_adj_by_year": inflation_adj_by_year
            }
            return

        years_list = list(range(1, years+1))
        fig = Figure(figsize=(9,3.2), dpi=95)
        ax = fig.add_subplot(111)
        ax.bar(years_list, fv_by_year, color=ACCENT_LINE)
        ax.set_facecolor(bg); fig.patch.set_facecolor(bg)
        ax.set_xlabel("Year"); ax.set_ylabel("Future Value (₹)")
        ax.set_title("Year-by-Year FV (SIP)", color=HEADING_FG)
        ax.tick_params(colors=TEXT_FG)
        for spine in ax.spines.values(): spine.set_color(TEXT_FG)
        ax.xaxis.label.set_color(TEXT_FG); ax.yaxis.label.set_color(TEXT_FG)
        ax.grid(axis="y", linestyle="--", alpha=0.3, color=ACCENT_LINE)
        canvas = FigureCanvasTkAgg(fig, master=chart_frame); canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        chart_canvas_container["canvas"] = canvas

        tree._calc = {
            "years": years,
            "sip": sip,
            "invested_by_year": invested_by_year,
            "fv_by_year": fv_by_year,
            "invested_total": invested_total,
            "fv_total": fv,
            "inflation_adj_total": inflation_adj_total,
            "inflation_adj_by_year": inflation_adj_by_year
        }

    def export_sip():
        data = getattr(tree, "_calc", None)
        if not data:
            messagebox.showinfo("No data","Calculate first."); return
        headers = ["Year","Monthly SIP","Invested","FV","InflationAdj FV"]
        rows = []
        for i in range(data["years"]):
            rows.append([
                i+1,
                round(data["sip"],2),
                round(data["invested_by_year"][i],2),
                round(data["fv_by_year"][i],2),
                round(data["inflation_adj_by_year"][i],2)
            ])
        rows.append([])
        rows.append(["Total Invested", data["invested_total"]])
        rows.append(["Final FV", data["fv_total"]])
        rows.append(["Inflation-adjusted FV", data["inflation_adj_total"]])
        path = save_to_excel_or_csv(
            os.path.join(REPORTS_DIR,f"SIP_Report_{today_str()}"),
            headers, rows
        )
        if path: messagebox.showinfo("Exported", f"Saved to {path}")

    calc_btn.config(command=calculate_sip)
    export_btn.config(command=export_sip)

# ---------- Loan Calculator ----------
def show_loan_calculator(frame):
    for w in frame.winfo_children(): w.destroy()
    bg = PRIMARY_BG
    frame.configure(bg=bg)

    tk.Label(frame, text="Loan Calculator", font=("Segoe UI", 18, "bold"),
             fg=HEADING_FG, bg=bg).pack(pady=(10,8))

    grid = tk.Frame(frame, bg=bg); grid.pack(fill="x", padx=16, pady=(0,8))
    tk.Label(grid, text="Loan Amount (₹):", bg=bg, fg=TEXT_FG).grid(row=0, column=0, sticky="w", padx=6, pady=6)
    e_amount = tk.Entry(grid, width=18); e_amount.grid(row=0, column=1, padx=6, pady=6)

    tk.Label(grid, text="Annual Interest Rate (%):", bg=bg, fg=TEXT_FG).grid(row=1, column=0, sticky="w", padx=6, pady=6)
    e_rate = tk.Entry(grid, width=18); e_rate.grid(row=1, column=1, padx=6, pady=6)

    tk.Label(grid, text="Tenure (years):", bg=bg, fg=TEXT_FG).grid(row=2, column=0, sticky="w", padx=6, pady=6)
    e_tenure = tk.Entry(grid, width=18); e_tenure.grid(row=2, column=1, padx=6, pady=6)

    btn_row = tk.Frame(frame, bg=bg); btn_row.pack(fill="x", padx=16, pady=(0,8))
    calc_btn = tk.Button(btn_row, text="Calculate EMI", bg=ACCENT_BTN, fg="black",
                         font=("Segoe UI",10,"bold"))
    export_btn = tk.Button(btn_row, text="Export Loan Report", bg=ACCENT_BTN_2, fg="black",
                           state="disabled", font=("Segoe UI",10,"bold"))
    calc_btn.pack(side="left", padx=(0,8)); export_btn.pack(side="left", padx=8)

    result_frame = tk.Frame(frame, bg=SIDEBAR_BG); result_frame.pack(fill="x", padx=16, pady=(6,8))
    result_label = tk.Label(result_frame, text="", bg=SIDEBAR_BG, fg="#FED7AA",
                            font=("Segoe UI",10), pady=6, justify="left", wraplength=1100)
    result_label.pack(fill="x")

    chart_frame = tk.Frame(frame, bg=bg); chart_frame.pack(fill="both", expand=True, padx=16, pady=(6,12))
    chart_canvas_container = {"canvas": None}

    def clear_chart():
        c = chart_canvas_container.get("canvas")
        if c:
            c.get_tk_widget().destroy()
            chart_canvas_container["canvas"] = None

    def calculate_loan():
        try:
            P = float(e_amount.get())
            annual_r = float(e_rate.get())
            years = int(e_tenure.get())
            if P <= 0 or annual_r <= 0 or years <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Invalid input","Enter positive numbers in all fields."); return

        n = years*12
        r = annual_r/12/100
        if r == 0:
            emi = P/n
        else:
            emi = P * r * (1+r)**n / ((1+r)**n - 1)
        total_payable = emi*n
        total_interest = total_payable - P

        result_label.config(text=(
            f"EMI: {format_currency(emi)}    Tenure: {years} years    "
            f"Total Interest: {format_currency(total_interest)}    "
            f"Total Payable: {format_currency(total_payable)}"
        ))

        export_btn.config(state="normal")

        clear_chart()
        if not MATPLOTLIB_AVAILABLE:
            tk.Label(chart_frame, text="Install matplotlib for chart.", fg=TEXT_FG, bg=bg).pack()
            chart_canvas_container["data"] = {"emi":emi,"n":n,"P":P,"interest":total_interest}
            return

        fig = Figure(figsize=(6,3.8), dpi=95)
        ax = fig.add_subplot(111)
        labels = ["Principal (₹)","Interest (₹)"]
        sizes = [P, total_interest]
        ax.set_facecolor(bg); fig.patch.set_facecolor(bg)
        wedges, texts, autotexts = ax.pie(
            sizes, labels=labels,
            autopct=lambda p: format_currency(p/100*sum(sizes)),
            startangle=90
        )
        for t in texts+autotexts: t.set_color(TEXT_FG)
        ax.axis("equal")
        ax.set_title("Principal vs Interest (Total over loan)", color=HEADING_FG)
        canvas = FigureCanvasTkAgg(fig, master=chart_frame); canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        chart_canvas_container["canvas"] = canvas
        chart_canvas_container["data"] = {"emi":emi,"n":n,"P":P,"interest":total_interest}

    def export_loan():
        data = chart_canvas_container.get("data")
        if not data:
            messagebox.showinfo("No data","Calculate EMI first."); return
        headers = ["Loan Amount","Annual Rate (%)","Tenure (years)","EMI (₹)","Total Interest (₹)","Total Payable (₹)"]
        rows = [[
            e_amount.get(), e_rate.get(), e_tenure.get(),
            round(data["emi"],2), round(data["interest"],2),
            round(data["emi"]*data["n"],2)
        ]]
        path = save_to_excel_or_csv(
            os.path.join(REPORTS_DIR,f"Loan_Report_{today_str()}"),
            headers, rows
        )
        if path: messagebox.showinfo("Exported", f"Saved to {path}")

    calc_btn.config(command=calculate_loan)
    export_btn.config(command=export_loan)

# ---------- FIRE Number Calculator ----------
# ---------- FIRE Number Calculator ----------
def show_fire_calculator(frame):
    for w in frame.winfo_children():
        w.destroy()

    bg = PRIMARY_BG
    frame.configure(bg=bg)

    tk.Label(
        frame,
        text="FIRE Calculator (How Much You Should Invest)",
        font=("Segoe UI", 18, "bold"),
        fg=HEADING_FG,
        bg=bg
    ).pack(pady=(10, 8))

    grid = tk.Frame(frame, bg=bg)
    grid.pack(fill="x", padx=16, pady=(0, 8))

    # ---- Inputs ----
    tk.Label(grid, text="Monthly Expenses (₹):", bg=bg, fg=TEXT_FG)\
        .grid(row=0, column=0, sticky="w", padx=6, pady=6)
    e_monthly = tk.Entry(grid, width=18)
    e_monthly.grid(row=0, column=1, padx=6, pady=6)

    tk.Label(grid, text="Current Savings (₹):", bg=bg, fg=TEXT_FG)\
        .grid(row=1, column=0, sticky="w", padx=6, pady=6)
    e_current = tk.Entry(grid, width=18)
    e_current.grid(row=1, column=1, padx=6, pady=6)

    tk.Label(grid, text="Years till Retirement:", bg=bg, fg=TEXT_FG)\
        .grid(row=2, column=0, sticky="w", padx=6, pady=6)
    e_years = tk.Entry(grid, width=18)
    e_years.grid(row=2, column=1, padx=6, pady=6)

    tk.Label(grid, text="Expected Annual Return (%):", bg=bg, fg=TEXT_FG)\
        .grid(row=3, column=0, sticky="w", padx=6, pady=6)
    e_return = tk.Entry(grid, width=18)
    e_return.grid(row=3, column=1, padx=6, pady=6)

    # ---- Buttons ----
    btn_row = tk.Frame(frame, bg=bg)
    btn_row.pack(fill="x", padx=16, pady=(0, 8))

    calc_btn = tk.Button(
        btn_row,
        text="Calculate Required Monthly Investment",
        bg=ACCENT_BTN_2,
        fg="black",
        font=("Segoe UI", 10, "bold")
    )
    export_btn = tk.Button(
        btn_row,
        text="Export FIRE Plan",
        bg=ACCENT_BTN,
        fg="black",
        state="disabled",
        font=("Segoe UI", 10, "bold")
    )
    calc_btn.pack(side="left", padx=(0, 8))
    export_btn.pack(side="left", padx=8)

    # ---- Result area ----
    result_frame = tk.Frame(frame, bg=SIDEBAR_BG)
    result_frame.pack(fill="x", padx=16, pady=(6, 8))

    result_label = tk.Label(
        result_frame,
        text="",
        bg=SIDEBAR_BG,
        fg="#BBF7D0",
        font=("Segoe UI", 10),
        pady=6,
        justify="left",
        wraplength=1100
    )
    result_label.pack(fill="x")

    # ---- Chart area ----
    chart_frame = tk.Frame(frame, bg=bg)
    chart_frame.pack(fill="both", expand=True, padx=16, pady=(6, 12))
    chart_canvas_container = {"canvas": None, "data": None}

    def clear_chart():
        c = chart_canvas_container.get("canvas")
        if c:
            c.get_tk_widget().destroy()
            chart_canvas_container["canvas"] = None

    def calculate_fire():
        # Read + validate
        try:
            monthly_exp = float(e_monthly.get())
            current = float(e_current.get())
            years = int(e_years.get())
            exp_return = float(e_return.get())
            if monthly_exp < 0 or current < 0 or years <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Invalid input", "Please enter valid positive numbers.")
            return

        # 1) FIRE target using 4% rule → 25× yearly expenses
        fire_target = monthly_exp * 12 * 25

        # 2) Required monthly SIP to reach FIRE in 'years'
        n_months = years * 12
        r_annual = exp_return / 100
        r_monthly = r_annual / 12 if r_annual != 0 else 0.0

        if current >= fire_target:
            required_monthly = 0.0
            proj_savings = [current] * years
        else:
            if r_monthly == 0:
                required_monthly = (fire_target - current) / n_months
            else:
                growth = (1 + r_monthly) ** n_months
                numerator = fire_target - current * growth
                denom = (growth - 1) / r_monthly
                if denom == 0:
                    messagebox.showerror("Error", "Please adjust inputs.")
                    return
                required_monthly = numerator / denom

            if required_monthly < 0:
                required_monthly = 0.0

            proj_savings = []
            bal = current
            for _y in range(years):
                for _ in range(12):
                    bal = bal * (1 + r_monthly) + required_monthly
                proj_savings.append(bal)

        # 3) Result text
        lines = [
            f"FIRE Target (25× yearly expenses): {format_currency(fire_target)}",
            f"Current Savings: {format_currency(current)}",
            f"Years till Retirement: {years}",
        ]
        if required_monthly == 0:
            lines.append("✅ You are already at or above your FIRE target.")
        else:
            lines.append(
                f"👉 You should invest about {format_currency(required_monthly)} per month "
                f"to reach FIRE in {years} year(s)."
            )
        lines.append(
            f"Projected corpus at retirement: {format_currency(proj_savings[-1])}"
        )
        result_label.config(text="\n".join(lines))

        # 4) Enable export
        export_btn.config(state="normal")

        # 5) Store for export / chart
        chart_canvas_container["data"] = {
            "proj": proj_savings,
            "years": years,
            "fire_target": fire_target,
            "required_monthly": required_monthly,
        }

        # 6) Chart
        clear_chart()
        if not MATPLOTLIB_AVAILABLE:
            tk.Label(
                chart_frame,
                text="Install matplotlib to view chart.",
                fg="white",
                bg=bg
            ).pack()
            return

        years_list = list(range(1, years + 1))
        fig = Figure(figsize=(9, 3.6), dpi=95)
        ax = fig.add_subplot(111)

        ax.plot(
            years_list,
            proj_savings,
            marker="o",
            label="Projected Savings",
            color=ACCENT_BTN
        )
        ax.hlines(
            fire_target,
            years_list[0],
            years_list[-1],
            colors="#FF7F50",
            linestyles="--",
            label="FIRE Target"
        )

        ax.set_xlabel("Year")
        ax.set_ylabel("Amount (₹)")
        ax.set_title("Your FIRE Journey", color=HEADING_FG)
        ax.grid(axis="y", linestyle="--", alpha=0.3, color=ACCENT_LINE)

        ax.set_facecolor(bg)
        fig.patch.set_facecolor(bg)
        ax.tick_params(colors=TEXT_FG)
        for spine in ax.spines.values():
            spine.set_color(TEXT_FG)
        ax.xaxis.label.set_color(TEXT_FG)
        ax.yaxis.label.set_color(TEXT_FG)
        ax.legend(facecolor=SIDEBAR_BG, edgecolor=TEXT_FG)

        canvas = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        chart_canvas_container["canvas"] = canvas

    def export_fire():
        data = chart_canvas_container.get("data")
        if not data:
            messagebox.showinfo("No data", "Please run the FIRE calculation first.")
            return

        headers = ["Year", "Projected Savings (₹)"]
        rows = []
        for i, val in enumerate(data["proj"]):
            rows.append([i + 1, round(val, 2)])
        rows.append([])
        rows.append(["Required Monthly Investment (₹)", round(data["required_monthly"], 2)])
        rows.append(["FIRE Target (₹)", round(data["fire_target"], 2)])

        path = save_to_excel_or_csv(
            os.path.join(REPORTS_DIR, f"FIRE_Plan_{today_str()}"),
            headers,
            rows
        )
        if path:
            messagebox.showinfo("Exported", f"Saved to {path}")

    calc_btn.config(command=calculate_fire)
    export_btn.config(command=export_fire)

# ---------- Inflation Impact Calculator ----------
def show_inflation_calculator(frame):
    for w in frame.winfo_children(): w.destroy()
    bg = PRIMARY_BG
    frame.configure(bg=bg)

    tk.Label(frame, text="Inflation Impact Calculator", font=("Segoe UI", 18, "bold"),
             fg=HEADING_FG, bg=bg).pack(pady=(10,8))

    grid = tk.Frame(frame, bg=bg); grid.pack(fill="x", padx=16, pady=(0,8))
    tk.Label(grid, text="Current Amount (₹):", bg=bg, fg=TEXT_FG).grid(row=0, column=0, sticky="w", padx=6, pady=6)
    e_amount = tk.Entry(grid, width=18); e_amount.grid(row=0, column=1, padx=6, pady=6)

    tk.Label(grid, text="Annual Inflation Rate (%):", bg=bg, fg=TEXT_FG).grid(row=1, column=0, sticky="w", padx=6, pady=6)
    e_rate = tk.Entry(grid, width=18); e_rate.grid(row=1, column=1, padx=6, pady=6)

    tk.Label(grid, text="Duration (years):", bg=bg, fg=TEXT_FG).grid(row=2, column=0, sticky="w", padx=6, pady=6)
    e_years = tk.Entry(grid, width=18); e_years.grid(row=2, column=1, padx=6, pady=6)

    btn_row = tk.Frame(frame, bg=bg); btn_row.pack(fill="x", padx=16, pady=(0,8))
    calc_btn = tk.Button(btn_row, text="Calculate Impact", bg=ACCENT_BTN, fg="black",
                         font=("Segoe UI",10,"bold"))
    export_btn = tk.Button(btn_row, text="Export Results", bg=ACCENT_BTN_2, fg="black",
                           state="disabled", font=("Segoe UI",10,"bold"))
    calc_btn.pack(side="left", padx=(0,8)); export_btn.pack(side="left", padx=8)

    table_frame = tk.Frame(frame, bg=bg); table_frame.pack(fill="both", expand=False, padx=16, pady=(6,8))
    cols = ("year","cum_infl","future_cost","purch_power")
    tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=8)
    heads = {
        "year":"Year","cum_infl":"Cumulative Inflation (%)",
        "future_cost":"Future Cost (₹)","purch_power":"Purchasing Power (₹)"
    }
    for c in cols: tree.heading(c, text=heads[c])
    tree.column("year", width=60, anchor="center")
    tree.column("cum_infl", width=170, anchor="e")
    tree.column("future_cost", width=150, anchor="e")
    tree.column("purch_power", width=180, anchor="e")
    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscroll=vsb.set); vsb.pack(side="right", fill="y"); tree.pack(side="left", fill="both", expand=True)

    result_frame = tk.Frame(frame, bg=SIDEBAR_BG); result_frame.pack(fill="x", padx=16, pady=(6,8))
    result_label = tk.Label(result_frame, text="", bg=SIDEBAR_BG, fg="#E9D5FF",
                            font=("Segoe UI",10), pady=6, justify="left", wraplength=1100)
    result_label.pack(fill="x")

    chart_frame = tk.Frame(frame, bg=bg); chart_frame.pack(fill="both", expand=True, padx=16, pady=(6,12))
    chart_canvas_container = {"canvas": None}

    def clear_chart():
        c = chart_canvas_container.get("canvas")
        if c:
            c.get_tk_widget().destroy()
            chart_canvas_container["canvas"] = None

    def calculate():
        try:
            amount = float(e_amount.get())
            rate = float(e_rate.get())
            years = int(e_years.get())
            if amount <= 0 or rate < 0 or years <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Invalid input","Enter valid positive numbers in all fields."); return

        r = rate/100
        future_costs = []; purch_power_list = []; cum_infl_list = []
        for y in range(1, years+1):
            factor = (1+r)**y
            cum_infl = factor - 1
            future_cost = amount * factor
            purch_power = amount / factor
            future_costs.append(future_cost)
            purch_power_list.append(purch_power)
            cum_infl_list.append(cum_infl)

        for rid in tree.get_children(): tree.delete(rid)
        for i in range(years):
            tree.insert("", "end", values=(
                i+1,
                f"{cum_infl_list[i]*100:.2f}%",
                format_currency(round(future_costs[i],2)),
                format_currency(round(purch_power_list[i],2))
            ))

        final_pp = purch_power_list[-1]; final_fc = future_costs[-1]
        result_label.config(text=(
            f"After {years} years at {rate:.2f}% inflation: "
            f"₹{amount:,.2f} will feel like {format_currency(final_pp)} today. "
            f"Future cost of the same item: {format_currency(final_fc)}."
        ))

        export_btn.config(state="normal")

        clear_chart()
        if not MATPLOTLIB_AVAILABLE:
            tk.Label(chart_frame, text="Install matplotlib for chart.", fg=TEXT_FG, bg=bg).pack()
            tree._calc = {
                "amount":amount,"rate":rate,"years":years,
                "future_costs":future_costs,
                "purch_power":purch_power_list,
                "cum_infl":cum_infl_list
            }
            return

        years_list = list(range(1, years+1))
        fig = Figure(figsize=(9,3.2), dpi=95)
        ax = fig.add_subplot(111)
        ax.plot(years_list, purch_power_list, marker="o", label="Purchasing Power", color=ACCENT_BTN)
        ax.plot(years_list, future_costs, marker="x", linestyle="--", label="Future Cost", color=ACCENT_LINE)
        ax.set_facecolor(bg); fig.patch.set_facecolor(bg)
        ax.set_xlabel("Year"); ax.set_ylabel("Amount (₹)")
        ax.set_title("Inflation Impact Over Time", color=HEADING_FG)
        ax.tick_params(colors=TEXT_FG)
        for spine in ax.spines.values(): spine.set_color(TEXT_FG)
        ax.xaxis.label.set_color(TEXT_FG); ax.yaxis.label.set_color(TEXT_FG)
        ax.grid(axis="y", linestyle="--", alpha=0.3, color=ACCENT_LINE)
        ax.legend(facecolor=SIDEBAR_BG, edgecolor=TEXT_FG)
        canvas = FigureCanvasTkAgg(fig, master=chart_frame); canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        chart_canvas_container["canvas"] = canvas

        tree._calc = {
            "amount":amount,"rate":rate,"years":years,
            "future_costs":future_costs,
            "purch_power":purch_power_list,
            "cum_infl":cum_infl_list
        }

    def export_results():
        data = getattr(tree, "_calc", None)
        if not data:
            messagebox.showinfo("No data","Calculate first."); return
        headers = ["Year","Cumulative Inflation (%)","Future Cost (₹)","Purchasing Power (₹)"]
        rows = []
        for i in range(data["years"]):
            rows.append([
                i+1,
                round(data["cum_infl"][i]*100,2),
                round(data["future_costs"][i],2),
                round(data["purch_power"][i],2)
            ])
        rows.append([]); rows.append(["Original Amount", data["amount"]])
        rows.append(["Inflation Rate (%)", data["rate"]])
        path = save_to_excel_or_csv(
            os.path.join(REPORTS_DIR,f"Inflation_Impact_{today_str()}"),
            headers, rows
        )
        if path: messagebox.showinfo("Exported", f"Saved to {path}")

    calc_btn.config(command=calculate)
    export_btn.config(command=export_results)

# ---------- Main UI ----------
root = tk.Tk()
root.title("💼 Personal Finance Toolkit")
root.geometry("1200x720")
root.configure(bg=SIDEBAR_BG)
root.minsize(1000, 650)

sidebar = tk.Frame(root, bg=SIDEBAR_BG, width=260)
sidebar.pack(side="left", fill="y")
main_frame = tk.Frame(root, bg=PRIMARY_BG)
main_frame.pack(side="right", fill="both", expand=True)

tk.Label(sidebar, text="💰 Finance Toolkit", fg=HEADING_FG, bg=SIDEBAR_BG,
         font=("Segoe UI",18,"bold")).pack(pady=18)

buttons = [
    ("SIP Calculator", lambda: show_sip_calculator(main_frame)),
    ("Step-up SIP vs SIP", lambda: show_step_up_vs_sip(main_frame)),
    ("FIRE Calculator", lambda: show_fire_calculator(main_frame)),
    ("Inflation Impact", lambda: show_inflation_calculator(main_frame)),
    ("Loan Calculator", lambda: show_loan_calculator(main_frame)),
    ("Expense Tracker", lambda: show_expense_tracker(main_frame)),
]

for text, cmd in buttons:
    tk.Button(
        sidebar,
        text=text,
        command=cmd,
        bg=SIDEBAR_BG,
        fg="white",
        font=("Segoe UI",11,"bold"),
        relief="flat",
        width=22,
        height=2,
        activebackground="#243B63",
        activeforeground="white",
        bd=0,
        highlightthickness=0
    ).pack(pady=4, padx=10)

tk.Label(sidebar, text="© Personal Finance Toolkit", fg="#94A3B8", bg=SIDEBAR_BG,
         font=("Segoe UI",9)).pack(side="bottom", pady=10)

show_sip_calculator(main_frame)
root.mainloop()